"""
Script para poblar la tabla SensorConfig con los datos del Excel Alertas_IdealControl.xlsm
Ejecutar: python poblar_sensores.py
"""
import os
import django
import openpyxl
import sys
import io

# Configurar salida UTF-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Configurar Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from incidencias.models import Centro, SensorConfig

def poblar_sensores():
    """Lee el Excel y crea los registros de SensorConfig"""
    
    archivo = 'Alertas_IdealControl.xlsm'
    
    print("Cargando Excel...")
    wb = openpyxl.load_workbook(archivo, data_only=True)
    
    # Usar la hoja "BASE DE DATOS" que tiene todos los sensores
    ws = wb['BASE DE DATOS']
    
    print(f"Procesando hoja 'BASE DE DATOS' con {ws.max_row} filas...")
    
    sensores_creados = 0
    sensores_actualizados = 0
    sensores_unicos = set()
    
    # Leer desde la fila 2 (fila 1 son encabezados)
    for row_idx in range(2, min(1000, ws.max_row + 1)):  # Limitar a primeras 1000 filas para evitar duplicados
        # Columnas: PISCICULTURA, SERVICIO, SISTEMA, EQUIPO, TIPO_MEDICION, LIMITE_MIN, LIMITE_MAX
        piscicultura = ws.cell(row=row_idx, column=1).value
        servicio = ws.cell(row=row_idx, column=2).value
        sistema = ws.cell(row=row_idx, column=3).value
        equipo = ws.cell(row=row_idx, column=4).value
        tipo_medicion = ws.cell(row=row_idx, column=5).value
        limite_min = ws.cell(row=row_idx, column=6).value
        limite_max = ws.cell(row=row_idx, column=7).value
        
        # Saltar filas vacías
        if not piscicultura or not sistema or not equipo:
            continue
        
        # Normalizar nombre del centro
        centro_nombre_map = {
            'LIQUIÑE': 'Liquiñe',
            'LIQUINE': 'Liquiñe',
            'LOS CIPRESES': 'Cipreses',
            'CIPRESES': 'Cipreses',
            'PCC': 'PCC',
            'RAHUE': 'Rahue',
            'SANTA JUANA': 'Santa Juana',
            'TRAFUN': 'Trafún',
            'TRAFÚN': 'Trafún',
            'ESPERANZA': 'Esperanza',
            'HUEYUSCA': 'Hueyusca'
        }
        
        centro_nombre = centro_nombre_map.get(piscicultura.upper().strip(), piscicultura)
        
        # Crear clave única para evitar duplicados
        clave_unica = f"{centro_nombre}|{sistema}|{equipo}"
        if clave_unica in sensores_unicos:
            continue
        
        # Buscar el centro
        try:
            centro = Centro.objects.get(nombre=centro_nombre)
        except Centro.DoesNotExist:
            # Intentar crear el centro si no existe
            try:
                centro = Centro.objects.create(
                    id=centro_nombre.lower().replace(' ', '-'),
                    nombre=centro_nombre
                )
                print(f"  [+] Centro creado: {centro_nombre}")
            except:
                print(f"  [!] Centro '{centro_nombre}' no se pudo crear, saltando...")
                continue
        
        # Limpiar valores
        sistema = sistema.strip() if sistema else ''
        equipo = equipo.strip() if equipo else ''
        tipo_medicion = tipo_medicion.strip() if tipo_medicion else ''
        limite_min = str(limite_min).strip() if limite_min and limite_min != 'None' else ''
        limite_max = str(limite_max).strip() if limite_max and limite_max != 'None' else ''
        
        # Crear o actualizar sensor
        sensor, created = SensorConfig.objects.update_or_create(
            centro=centro,
            sistema=sistema,
            equipo=equipo,
            defaults={
                'tipo_medicion': tipo_medicion,
                'limite_min': limite_min,
                'limite_max': limite_max,
                'activo': True,
                'orden': len(sensores_unicos)
            }
        )
        
        sensores_unicos.add(clave_unica)
        
        if created:
            sensores_creados += 1
            if sensores_creados % 10 == 0:
                print(f"  [+] {sensores_creados} sensores creados...")
        else:
            sensores_actualizados += 1
    
    print(f"\n{'='*80}")
    print(f"RESUMEN:")
    print(f"  - Sensores creados: {sensores_creados}")
    print(f"  - Sensores actualizados: {sensores_actualizados}")
    print(f"  - Total procesados: {sensores_creados + sensores_actualizados}")
    print(f"{'='*80}\n")

if __name__ == '__main__':
    print("="*80)
    print("POBLANDO TABLA SensorConfig DESDE EXCEL")
    print("="*80)
    poblar_sensores()
    print("[OK] Proceso completado")
