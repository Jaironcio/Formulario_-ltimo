"""
Script para analizar el archivo Alertas_IdealControl.xlsm
y extraer su estructura para diseñar el formulario de sensores
"""
import openpyxl
import json
from datetime import datetime
import sys
import io

# Configurar salida UTF-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analizar_excel():
    archivo = 'Alertas_IdealControl.xlsm'
    
    try:
        print("Cargando archivo Excel...")
        wb = openpyxl.load_workbook(archivo, data_only=True)
        
        print("=" * 100)
        print("ANALISIS DEL ARCHIVO: Alertas_IdealControl.xlsm")
        print("=" * 100)
        
        # Listar todas las hojas
        print(f"\nHOJAS DISPONIBLES ({len(wb.sheetnames)}):")
        for i, nombre in enumerate(wb.sheetnames, 1):
            print(f"  {i}. {nombre}")
        
        resultados = {
            'fecha_analisis': datetime.now().isoformat(),
            'hojas': [],
            'estructura_por_centro': {}
        }
        
        # Analizar cada hoja
        for nombre_hoja in wb.sheetnames:
            print(f"\n{'=' * 100}")
            print(f"HOJA: {nombre_hoja}")
            print(f"{'=' * 100}")
            
            ws = wb[nombre_hoja]
            
            # Dimensiones
            print(f"\nDIMENSIONES:")
            print(f"  - Filas con datos: {ws.max_row}")
            print(f"  - Columnas con datos: {ws.max_column}")
            
            # Extraer encabezados
            encabezados = []
            print(f"\nENCABEZADOS:")
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col_idx).value
                encabezados.append(header)
                if header:
                    print(f"  Col {col_idx:2d}: {header}")
            
            # Mostrar primeras 30 filas
            print(f"\nPRIMERAS 30 FILAS DE DATOS:")
            print("-" * 100)
            
            datos_hoja = []
            for row_idx in range(1, min(31, ws.max_row + 1)):
                row_data = []
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    row_data.append(value)
                
                datos_hoja.append(row_data)
                
                # Imprimir fila formateada
                row_str = " | ".join([
                    str(v)[:20] if v is not None else "" 
                    for v in row_data[:10]  # Primeras 10 columnas
                ])
                print(f"Fila {row_idx:2d}: {row_str}")
            
            # Guardar en resultados
            resultados['hojas'].append({
                'nombre': nombre_hoja,
                'filas': ws.max_row,
                'columnas': ws.max_column,
                'encabezados': encabezados,
                'primeras_filas': datos_hoja
            })
        
        # Guardar análisis en archivo JSON
        print(f"\n{'=' * 100}")
        print("Guardando analisis en JSON...")
        
        with open('analisis_excel_sensores.json', 'w', encoding='utf-8') as f:
            json.dump(resultados, f, indent=2, ensure_ascii=False, default=str)
        
        print("EXITO: Analisis completado exitosamente")
        print("Archivo guardado: analisis_excel_sensores.json")
        print(f"{'=' * 100}\n")
        
        # Resumen
        print("RESUMEN:")
        print(f"  - Total de hojas analizadas: {len(resultados['hojas'])}")
        for hoja in resultados['hojas']:
            print(f"  - {hoja['nombre']}: {hoja['filas']} filas x {hoja['columnas']} columnas")
        
    except FileNotFoundError:
        print(f"ERROR: No se encontro el archivo '{archivo}'")
        print("   Asegurate de que el archivo este en el mismo directorio que este script")
    except Exception as e:
        print(f"ERROR al analizar el archivo: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    analizar_excel()
