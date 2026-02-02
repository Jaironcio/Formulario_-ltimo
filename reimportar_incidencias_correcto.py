"""
Script para reimportar incidencias correctamente desde el Excel
preservando EXACTAMENTE los tipos de incidencia tal como están
"""
import os
import django
import openpyxl
from datetime import datetime, time

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from incidencias.models import Centro, Operario, Incidencia

# ELIMINAR TODAS LAS INCIDENCIAS EXISTENTES
print("Eliminando incidencias existentes...")
Incidencia.objects.all().delete()
print(f"Incidencias eliminadas. Total en BD: {Incidencia.objects.count()}")

def limpiar_valor(valor):
    if valor is None:
        return ''
    if isinstance(valor, str):
        return valor.strip()
    return str(valor)

def parsear_fecha_hora(fecha_val, hora_val):
    try:
        if isinstance(fecha_val, datetime):
            fecha = fecha_val.date()
        else:
            fecha = datetime.strptime(str(fecha_val), '%Y-%m-%d %H:%M:%S').date()
        
        if isinstance(hora_val, time):
            hora = hora_val
        elif isinstance(hora_val, datetime):
            hora = hora_val.time()
        else:
            hora = datetime.strptime(str(hora_val), '%H:%M:%S').time()
        
        return datetime.combine(fecha, hora)
    except:
        if isinstance(fecha_val, datetime):
            return fecha_val
        return datetime.strptime(str(fecha_val).split()[0], '%Y-%m-%d')

def normalizar_centro(nombre_centro):
    mapeo = {
        'TRAFUN': 'trafun',
        'TRAFÚN': 'trafun',
        'LIQUIÑE': 'liquine',
        'LIQUINE': 'liquine',
        'CIPRESES': 'cipreses',
        'LOS CIPRESES': 'cipreses',
        'PCC': 'pcc',
        'RAHUE': 'rahue',
        'SANTA JUANA': 'santa-juana',
        'ESPERANZA': 'esperanza',
        'HUEYUSCA': 'hueyusca'
    }
    nombre_upper = nombre_centro.upper().strip()
    return mapeo.get(nombre_upper, nombre_centro.lower().replace(' ', '-'))

def normalizar_turno(turno_excel):
    turno = turno_excel.upper().strip()
    if 'NOCHE' in turno:
        return 'Noche'
    elif 'TARDE' in turno:
        return 'Tarde'
    elif 'DIA' in turno or 'MAÑANA' in turno or 'MA' in turno:
        return 'Mañana'
    return turno

def normalizar_si_no(valor):
    if not valor:
        return False
    valor_str = str(valor).strip().upper()
    return valor_str in ['SI', 'SÍ', 'S', 'YES', 'Y', '1', 'TRUE']

def buscar_operario(nombre_persona, cargo, centro_id):
    if not nombre_persona:
        return None
    
    try:
        centro = Centro.objects.get(id=centro_id)
        operario = Operario.objects.filter(
            centro=centro,
            nombre__icontains=nombre_persona.strip()
        ).first()
        
        if operario:
            return operario
        
        if cargo:
            operario = Operario.objects.filter(
                centro=centro,
                cargo__icontains=cargo.strip()
            ).first()
            if operario:
                return operario
        
        return Operario.objects.filter(centro=centro).first()
    except:
        return None

def importar_hoja(ws, nombre_hoja):
    print(f"\n{'='*80}")
    print(f"Importando hoja: {nombre_hoja}")
    print(f"{'='*80}")
    
    importadas = 0
    errores = 0
    
    for row_idx in range(2, ws.max_row + 1):
        fecha_val = ws.cell(row=row_idx, column=1).value
        if not fecha_val:
            continue
        
        try:
            fecha_val = ws.cell(row=row_idx, column=1).value
            hora_val = ws.cell(row=row_idx, column=2).value
            turno = limpiar_valor(ws.cell(row=row_idx, column=3).value)
            centro_nombre = limpiar_valor(ws.cell(row=row_idx, column=4).value)
            modulo = limpiar_valor(ws.cell(row=row_idx, column=5).value)
            estanque = limpiar_valor(ws.cell(row=row_idx, column=6).value)
            tipo_incidencia = limpiar_valor(ws.cell(row=row_idx, column=7).value)
            sensor = limpiar_valor(ws.cell(row=row_idx, column=8).value)
            proveedor = limpiar_valor(ws.cell(row=row_idx, column=9).value)
            valor = limpiar_valor(ws.cell(row=row_idx, column=10).value)
            tiempo_resolucion = ws.cell(row=row_idx, column=11).value
            riesgo_peces = limpiar_valor(ws.cell(row=row_idx, column=12).value)
            perdida_economica = limpiar_valor(ws.cell(row=row_idx, column=13).value)
            riesgo_personas = limpiar_valor(ws.cell(row=row_idx, column=14).value)
            
            # COLUMNA 15 - ESTE ES EL TIPO DE INCIDENCIA NORMALIZADA CORRECTO
            incidencia_normalizada = limpiar_valor(ws.cell(row=row_idx, column=15).value)
            
            observacion = limpiar_valor(ws.cell(row=row_idx, column=16).value)
            persona_llamada = limpiar_valor(ws.cell(row=row_idx, column=17).value)
            cargo = limpiar_valor(ws.cell(row=row_idx, column=18).value)
            
            if not centro_nombre:
                continue
            
            fecha_hora = parsear_fecha_hora(fecha_val, hora_val)
            centro_id = normalizar_centro(centro_nombre)
            
            try:
                centro = Centro.objects.get(id=centro_id)
            except Centro.DoesNotExist:
                errores += 1
                continue
            
            turno_normalizado = normalizar_turno(turno)
            operario = buscar_operario(persona_llamada, cargo, centro_id)
            
            try:
                tiempo_res = int(tiempo_resolucion) if tiempo_resolucion else None
            except:
                tiempo_res = None
            
            # Determinar parámetros afectados basado en el sensor
            parametros = ''
            oxigeno_nivel = ''
            oxigeno_valor = ''
            temperatura_nivel = ''
            temperatura_valor = ''
            
            if sensor:
                sensor_lower = sensor.lower()
                if 'oxigeno' in sensor_lower or 'oxígeno' in sensor_lower:
                    parametros = 'oxigeno'
                    if 'alto' in tipo_incidencia.lower():
                        oxigeno_nivel = 'alta'
                    elif 'bajo' in tipo_incidencia.lower():
                        oxigeno_nivel = 'baja'
                    oxigeno_valor = valor
                elif 'temperatura' in sensor_lower:
                    parametros = 'temperatura'
                    if 'alta' in tipo_incidencia.lower():
                        temperatura_nivel = 'alta'
                    elif 'baja' in tipo_incidencia.lower():
                        temperatura_nivel = 'baja'
                    temperatura_valor = valor
            
            # Crear incidencia con el tipo EXACTO del Excel
            incidencia = Incidencia.objects.create(
                fecha_hora=fecha_hora,
                turno=turno_normalizado,
                centro=centro,
                tipo_incidencia='modulos',
                modulo=modulo,
                estanque=estanque,
                parametros_afectados=parametros,
                oxigeno_nivel=oxigeno_nivel,
                oxigeno_valor=oxigeno_valor,
                temperatura_nivel=temperatura_nivel,
                temperatura_valor=temperatura_valor,
                plataforma=proveedor if proveedor else '',
                tiempo_resolucion=tiempo_res,
                riesgo_peces=normalizar_si_no(riesgo_peces),
                perdida_economica=normalizar_si_no(perdida_economica),
                riesgo_personas=normalizar_si_no(riesgo_personas),
                observacion=observacion,
                operario_contacto=operario,
                # USAR EL VALOR EXACTO DE LA COLUMNA 15 DEL EXCEL
                tipo_incidencia_normalizada=incidencia_normalizada if incidencia_normalizada else tipo_incidencia
            )
            
            importadas += 1
            if importadas % 100 == 0:
                print(f"  [+] {importadas} incidencias importadas...")
        
        except Exception as e:
            errores += 1
            if errores <= 3:
                print(f"  [ERROR] Fila {row_idx}: {str(e)}")
    
    print(f"\n  Resumen hoja '{nombre_hoja}':")
    print(f"    - Importadas: {importadas}")
    print(f"    - Errores: {errores}")
    
    return importadas, errores

print("="*80)
print("REIMPORTANDO INCIDENCIAS DESDE Incidencias_Completo.xlsm")
print("="*80)

archivo = 'Incidencias_Completo.xlsm'
wb = openpyxl.load_workbook(archivo, data_only=True)

total_importadas = 0
total_errores = 0

hojas_turnos = ['TURNO NOCHE', 'TURNO TARDE', 'TURNO MAÑANA']

for nombre_hoja in hojas_turnos:
    if nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        importadas, errores = importar_hoja(ws, nombre_hoja)
        total_importadas += importadas
        total_errores += errores

print(f"\n{'='*80}")
print("RESUMEN FINAL")
print(f"{'='*80}")
print(f"Total incidencias importadas: {total_importadas}")
print(f"Total errores: {total_errores}")
print(f"Incidencias en BD: {Incidencia.objects.count()}")
print(f"{'='*80}\n")

# Mostrar tipos únicos
from django.db.models import Count
print("\nTIPOS DE INCIDENCIA IMPORTADOS:")
print("-"*80)
tipos = Incidencia.objects.values('tipo_incidencia_normalizada').annotate(
    total=Count('id')
).order_by('-total')

for item in tipos:
    print(f"  {item['tipo_incidencia_normalizada']:50} {item['total']:4}")
