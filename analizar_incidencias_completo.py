"""
Script para analizar la estructura del Excel Incidencias_Completo.xlsm
"""
import openpyxl
import json

archivo = 'Incidencias_Completo.xlsm'

print("="*80)
print("ANALIZANDO ESTRUCTURA DE Incidencias_Completo.xlsm")
print("="*80)

wb = openpyxl.load_workbook(archivo, data_only=True)

print(f"\nHojas disponibles: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"HOJA: {sheet_name}")
    print(f"{'='*80}")
    print(f"Dimensiones: {ws.max_row} filas x {ws.max_column} columnas")
    
    # Leer encabezados (primera fila)
    print("\nEncabezados:")
    headers = []
    for col in range(1, min(ws.max_column + 1, 30)):
        header = ws.cell(row=1, column=col).value
        if header:
            headers.append(header)
            print(f"  Col {col}: {header}")
    
    # Mostrar primeras 3 filas de datos
    print("\nPrimeras 3 filas de datos:")
    for row_idx in range(2, min(5, ws.max_row + 1)):
        print(f"\n  Fila {row_idx}:")
        for col_idx, header in enumerate(headers, start=1):
            valor = ws.cell(row=row_idx, column=col_idx).value
            if valor:
                print(f"    {header}: {valor}")

print("\n" + "="*80)
print("ANALISIS COMPLETADO")
print("="*80)
