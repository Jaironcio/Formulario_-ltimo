"""
Analizar qué registros tienen la columna 15 vacía en el Excel
"""
import openpyxl

archivo = 'Incidencias_Completo.xlsm'
wb = openpyxl.load_workbook(archivo, data_only=True)

print("="*80)
print("ANALIZANDO REGISTROS CON COLUMNA 15 VACIA")
print("="*80)

hojas = ['TURNO NOCHE', 'TURNO TARDE', 'TURNO MAÑANA']
total_vacios = 0
total_registros = 0

for nombre_hoja in hojas:
    if nombre_hoja not in wb.sheetnames:
        continue
    
    ws = wb[nombre_hoja]
    vacios_hoja = 0
    registros_hoja = 0
    
    print(f"\n{'='*80}")
    print(f"Hoja: {nombre_hoja}")
    print(f"{'='*80}")
    
    print("\nPrimeros 10 registros con columna 15 vacía:")
    print(f"{'Fila':<6} {'Fecha':<12} {'Centro':<15} {'Módulo':<15} {'Col7:TipoInc':<20} {'Col15:Incidencia':<20}")
    print("-"*80)
    
    mostrados = 0
    
    for row_idx in range(2, ws.max_row + 1):
        fecha = ws.cell(row=row_idx, column=1).value
        if not fecha:
            continue
        
        registros_hoja += 1
        col15 = ws.cell(row=row_idx, column=15).value
        
        if not col15 or str(col15).strip() == '':
            vacios_hoja += 1
            
            if mostrados < 10:
                centro = ws.cell(row=row_idx, column=4).value
                modulo = ws.cell(row=row_idx, column=5).value
                col7 = ws.cell(row=row_idx, column=7).value
                
                fecha_str = str(fecha)[:10] if fecha else ''
                centro_str = str(centro)[:13] if centro else ''
                modulo_str = str(modulo)[:13] if modulo else ''
                col7_str = str(col7)[:18] if col7 else ''
                col15_str = str(col15)[:18] if col15 else '(vacío)'
                
                print(f"{row_idx:<6} {fecha_str:<12} {centro_str:<15} {modulo_str:<15} {col7_str:<20} {col15_str:<20}")
                mostrados += 1
    
    print(f"\nResumen {nombre_hoja}:")
    print(f"  Total registros: {registros_hoja}")
    print(f"  Con columna 15 vacía: {vacios_hoja}")
    print(f"  Con columna 15 llena: {registros_hoja - vacios_hoja}")
    
    total_vacios += vacios_hoja
    total_registros += registros_hoja

print(f"\n{'='*80}")
print("RESUMEN GENERAL")
print(f"{'='*80}")
print(f"Total de registros en Excel: {total_registros}")
print(f"Registros con columna 15 vacía: {total_vacios}")
print(f"Registros con columna 15 llena: {total_registros - total_vacios}")
print(f"\nEstos {total_vacios} registros usan la columna 7 como tipo de incidencia")
