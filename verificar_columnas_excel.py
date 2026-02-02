"""
Verificar qué columna tiene los tipos de incidencia correctos
"""
import openpyxl

archivo = 'Incidencias_Completo.xlsm'
wb = openpyxl.load_workbook(archivo, data_only=True)

# Revisar TURNO TARDE (tiene más variedad)
ws = wb['TURNO TARDE']

print("="*80)
print("COMPARANDO COLUMNAS 7 (TipoIncidencia) vs 15 (Incidencias)")
print("="*80)

print("\nPrimeras 10 filas de ejemplo:")
print("-"*80)
print(f"{'Fila':<6} {'Col 7: TipoIncidencia':<40} {'Col 15: Incidencias':<40}")
print("-"*80)

for row_idx in range(2, 12):
    col7 = ws.cell(row=row_idx, column=7).value
    col15 = ws.cell(row=row_idx, column=15).value
    print(f"{row_idx:<6} {str(col7)[:38]:<40} {str(col15)[:38]:<40}")

# Contar valores únicos en cada columna
from collections import Counter

col7_valores = []
col15_valores = []

for row_idx in range(2, ws.max_row + 1):
    val7 = ws.cell(row=row_idx, column=7).value
    val15 = ws.cell(row=row_idx, column=15).value
    
    if val7:
        col7_valores.append(str(val7).strip())
    if val15:
        col15_valores.append(str(val15).strip())

print("\n" + "="*80)
print("VALORES UNICOS EN COLUMNA 7 (TipoIncidencia):")
print("="*80)
counter7 = Counter(col7_valores)
for val, count in counter7.most_common():
    print(f"  {val:<50} {count:4}")

print("\n" + "="*80)
print("VALORES UNICOS EN COLUMNA 15 (Incidencias):")
print("="*80)
counter15 = Counter(col15_valores)
for val, count in counter15.most_common():
    print(f"  {val:<50} {count:4}")
