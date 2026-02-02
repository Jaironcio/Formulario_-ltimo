"""
Script para analizar los tipos de incidencia únicos en el Excel
"""
import openpyxl
from collections import Counter

archivo = 'Incidencias_Completo.xlsm'

print("="*80)
print("ANALIZANDO TIPOS DE INCIDENCIA EN EL EXCEL")
print("="*80)

wb = openpyxl.load_workbook(archivo, data_only=True)

todos_tipos = []

# Analizar cada hoja de turno
hojas = ['TURNO NOCHE', 'TURNO TARDE', 'TURNO MAÑANA']

for nombre_hoja in hojas:
    if nombre_hoja not in wb.sheetnames:
        continue
    
    ws = wb[nombre_hoja]
    print(f"\n{'='*80}")
    print(f"Hoja: {nombre_hoja}")
    print(f"{'='*80}")
    
    # La columna 15 es "Incidencia" o "Incidencias"
    tipos_hoja = []
    
    for row_idx in range(2, ws.max_row + 1):
        tipo = ws.cell(row=row_idx, column=15).value
        if tipo and str(tipo).strip():
            tipo_limpio = str(tipo).strip()
            tipos_hoja.append(tipo_limpio)
            todos_tipos.append(tipo_limpio)
    
    # Contar tipos únicos en esta hoja
    counter = Counter(tipos_hoja)
    print(f"\nTipos únicos encontrados: {len(counter)}")
    for tipo, cantidad in counter.most_common():
        print(f"  {tipo:50} {cantidad:4} veces")

# Resumen general
print(f"\n{'='*80}")
print("RESUMEN GENERAL - TODOS LOS TIPOS DE INCIDENCIA")
print(f"{'='*80}")

counter_total = Counter(todos_tipos)
print(f"\nTotal de tipos únicos: {len(counter_total)}")
print(f"Total de registros: {len(todos_tipos)}\n")

for tipo, cantidad in counter_total.most_common():
    porcentaje = (cantidad / len(todos_tipos)) * 100
    print(f"  {tipo:50} {cantidad:4} ({porcentaje:5.1f}%)")

print(f"\n{'='*80}")
print("LISTA PARA COPIAR (formato Python):")
print(f"{'='*80}\n")

print("tipos_incidencia = [")
for tipo, _ in counter_total.most_common():
    print(f"    '{tipo}',")
print("]")
